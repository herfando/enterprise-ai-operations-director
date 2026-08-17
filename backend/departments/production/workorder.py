from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from database_snowflake.connection import get_snowflake_connection

router = APIRouter(
    prefix="/production",
    tags=["Production Work Order"],
)


# =====================================================
# GENERATE WORK ORDER ID
# =====================================================


def generate_work_order_id():
    """
    Generate unique Work Order ID.

    Example:
        WO-PROD-20260817-123456
    """

    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    return f"WO-PROD-{timestamp}"


# =====================================================
# GET MAINTENANCE REQUEST
# =====================================================


def get_maintenance_request(request_id: str):
    """
    Get an existing Maintenance Request from Snowflake.
    """

    conn = get_snowflake_connection()
    cursor = conn.cursor()

    try:

        query = """
            SELECT
                REQUEST_ID,
                REQUEST_DATE,
                REQUESTER_DEPARTMENT,
                MAINTENANCE_DEPARTMENT,
                PRODUCTION_START_DATE,
                PRODUCTION_END_DATE,
                MACHINE_NAME,
                PRODUCT_NAME,
                MAINTENANCE_TYPE,
                SEVERITY,
                PRIORITY,
                TOTAL_DOWNTIME_MINUTES,
                TOTAL_PLANNING,
                TOTAL_PRODUCTION,
                GOOD_PRODUCT,
                REJECT_PRODUCT,
                RECORD_COUNT,
                PROBLEM_DESCRIPTION,
                STATUS

            FROM DATABASE_SNOWFLAKE.MASTER_DATA.MAINTENANCE_REQUEST

            WHERE REQUEST_ID = %s
        """

        cursor.execute(
            query,
            (request_id,),
        )

        row = cursor.fetchone()

        if not row:
            return None

        return {
            "request_id": row[0],
            "request_date": row[1],
            "requester_department": row[2],
            "maintenance_department": row[3],
            "production_start_date": row[4],
            "production_end_date": row[5],
            "machine_name": row[6],
            "product_name": row[7],
            "maintenance_type": row[8],
            "severity": row[9],
            "priority": row[10],
            "total_downtime_minutes": row[11],
            "total_planning": row[12],
            "total_production": row[13],
            "good_product": row[14],
            "reject_product": row[15],
            "record_count": row[16],
            "problem_description": row[17],
            "status": row[18],
        }

    finally:

        cursor.close()
        conn.close()


# =====================================================
# BUILD WORK ORDER
# =====================================================


def build_work_order(request):
    """
    Build Work Order data from Maintenance Request.

    Production information is automatically populated.
    Maintenance execution fields remain empty.
    """

    work_order_id = generate_work_order_id()

    total_planning = request["total_planning"] or 0
    total_production = request["total_production"] or 0

    production_achievement = 0

    if total_planning > 0:

        production_achievement = (total_production / total_planning) * 100

    return {
        # =================================================
        # WO INFORMATION
        # =================================================
        "work_order_id": work_order_id,
        "request_id": request["request_id"],
        "wo_date": datetime.now().strftime("%Y-%m-%d"),
        "status": "Open",
        "priority": request["priority"],
        "maintenance_type": request["maintenance_type"],
        "requester_department": request["requester_department"],
        "maintenance_department": request["maintenance_department"],
        # =================================================
        # EQUIPMENT & PROBLEM
        # =================================================
        "machine_name": request["machine_name"],
        "product_name": request["product_name"],
        "production_period": (
            f"{request['production_start_date']} "
            f"to "
            f"{request['production_end_date']}"
        ),
        "total_downtime_minutes": request["total_downtime_minutes"],
        "production_achievement": round(
            production_achievement,
            2,
        ),
        "reject_product": request["reject_product"],
        "problem_description": request["problem_description"],
        "production_impact": (
            f"Production achieved "
            f"{round(production_achievement, 2)}% "
            f"of planned production."
        ),
        "requested_action": (
            "Inspect and troubleshoot the machine "
            "to identify the cause of production "
            "downtime and restore normal operation."
        ),
        # =================================================
        # MAINTENANCE EXECUTION
        #
        # EMPTY - FILLED BY MAINTENANCE
        # =================================================
        "maintenance_pic": "",
        "maintenance_start": "",
        "maintenance_finish": "",
        "failure_damage": "",
        "root_cause": "",
        "repair_action": "",
        "spare_part_used": "",
        "spare_part_quantity": "",
        # =================================================
        # CLOSURE
        #
        # EMPTY - FILLED DURING MAINTENANCE PROCESS
        # =================================================
        "machine_status": "",
        "verification": "",
        "remarks": "",
        "maintenance_supervisor": "",
    }


# =====================================================
# CREATE EXCEL WORK ORDER
# =====================================================


def create_work_order_excel(work_order):
    """
    Generate Excel Work Order document.
    """

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Maintenance Work Order"

    # =================================================
    # STYLES
    # =================================================

    title_font = Font(
        bold=True,
        size=16,
    )

    section_font = Font(
        bold=True,
        size=12,
    )

    label_font = Font(
        bold=True,
    )

    thin_side = Side(
        style="thin",
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    # =================================================
    # TITLE
    # =================================================

    worksheet.merge_cells(
        "A1:B1",
    )

    worksheet["A1"] = "MAINTENANCE WORK ORDER"

    worksheet["A1"].font = title_font

    worksheet["A1"].alignment = Alignment(
        horizontal="center",
    )

    # =================================================
    # WO INFORMATION
    # =================================================

    worksheet.merge_cells(
        "A3:B3",
    )

    worksheet["A3"] = "WO INFORMATION"

    worksheet["A3"].font = section_font

    wo_information = [
        ("Work Order ID", work_order["work_order_id"]),
        ("Request ID", work_order["request_id"]),
        ("WO Date", work_order["wo_date"]),
        ("Status", work_order["status"]),
        ("Priority", work_order["priority"]),
        ("Maintenance Type", work_order["maintenance_type"]),
        (
            "Requester Department",
            work_order["requester_department"],
        ),
        (
            "Maintenance Department",
            work_order["maintenance_department"],
        ),
    ]

    current_row = 4

    for label, value in wo_information:

        worksheet.cell(
            current_row,
            1,
            label,
        )

        worksheet.cell(
            current_row,
            2,
            value,
        )

        worksheet.cell(
            current_row,
            1,
        ).font = label_font

        current_row += 1

    # =================================================
    # EQUIPMENT & PROBLEM
    # =================================================

    current_row += 1

    worksheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=2,
    )

    worksheet.cell(
        current_row,
        1,
        "EQUIPMENT & PROBLEM",
    )

    worksheet.cell(
        current_row,
        1,
    ).font = section_font

    current_row += 1

    equipment_problem = [
        ("Machine", work_order["machine_name"]),
        ("Product", work_order["product_name"]),
        (
            "Production Period",
            work_order["production_period"],
        ),
        (
            "Total Downtime (minutes)",
            work_order["total_downtime_minutes"],
        ),
        (
            "Production Achievement (%)",
            work_order["production_achievement"],
        ),
        (
            "Reject Product (pcs)",
            work_order["reject_product"],
        ),
        (
            "Problem Description",
            work_order["problem_description"],
        ),
        (
            "Production Impact",
            work_order["production_impact"],
        ),
        (
            "Requested Action",
            work_order["requested_action"],
        ),
    ]

    for label, value in equipment_problem:

        worksheet.cell(
            current_row,
            1,
            label,
        )

        worksheet.cell(
            current_row,
            2,
            value,
        )

        worksheet.cell(
            current_row,
            1,
        ).font = label_font

        current_row += 1

    # =================================================
    # MAINTENANCE EXECUTION
    # =================================================

    current_row += 1

    worksheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=2,
    )

    worksheet.cell(
        current_row,
        1,
        "MAINTENANCE EXECUTION",
    )

    worksheet.cell(
        current_row,
        1,
    ).font = section_font

    current_row += 1

    maintenance_fields = [
        ("Maintenance PIC", work_order["maintenance_pic"]),
        ("Maintenance Start", work_order["maintenance_start"]),
        ("Maintenance Finish", work_order["maintenance_finish"]),
        (
            "Failure / Damage",
            work_order["failure_damage"],
        ),
        ("Root Cause", work_order["root_cause"]),
        ("Repair Action", work_order["repair_action"]),
        (
            "Spare Part Used",
            work_order["spare_part_used"],
        ),
        (
            "Spare Part Quantity",
            work_order["spare_part_quantity"],
        ),
    ]

    for label, value in maintenance_fields:

        worksheet.cell(
            current_row,
            1,
            label,
        )

        worksheet.cell(
            current_row,
            2,
            value,
        )

        worksheet.cell(
            current_row,
            1,
        ).font = label_font

        current_row += 1

    # =================================================
    # CLOSURE
    # =================================================

    current_row += 1

    worksheet.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=2,
    )

    worksheet.cell(
        current_row,
        1,
        "VERIFICATION & CLOSURE",
    )

    worksheet.cell(
        current_row,
        1,
    ).font = section_font

    current_row += 1

    closure_fields = [
        (
            "Machine Status",
            work_order["machine_status"],
        ),
        (
            "Verification",
            work_order["verification"],
        ),
        (
            "Remarks",
            work_order["remarks"],
        ),
        (
            "Maintenance Supervisor",
            work_order["maintenance_supervisor"],
        ),
    ]

    for label, value in closure_fields:

        worksheet.cell(
            current_row,
            1,
            label,
        )

        worksheet.cell(
            current_row,
            2,
            value,
        )

        worksheet.cell(
            current_row,
            1,
        ).font = label_font

        current_row += 1

    # =================================================
    # BORDER & ALIGNMENT
    # =================================================

    for row in worksheet.iter_rows(
        min_row=3,
        max_row=current_row - 1,
        min_col=1,
        max_col=2,
    ):

        for cell in row:

            cell.border = border

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # =================================================
    # COLUMN WIDTH
    # =================================================

    worksheet.column_dimensions["A"].width = 30

    worksheet.column_dimensions["B"].width = 65

    # =================================================
    # FREEZE
    # =================================================

    worksheet.freeze_panes = "A4"

    # =================================================
    # SAVE TO MEMORY
    # =================================================

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# =====================================================
# CREATE PDF WORK ORDER
# =====================================================


def create_work_order_pdf(work_order):
    """
    Generate PDF Maintenance Work Order.
    """

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()

    elements = []

    # =================================================
    # TITLE
    # =================================================

    elements.append(
        Paragraph(
            "<b>MAINTENANCE WORK ORDER</b>",
            styles["Title"],
        )
    )

    elements.append(Spacer(1, 15))

    # =================================================
    # HELPER
    # =================================================

    def create_section(title, fields):

        data = [
            [
                Paragraph(
                    f"<b>{title}</b>",
                    styles["Heading3"],
                ),
                "",
            ]
        ]

        for label, value in fields:

            data.append(
                [
                    Paragraph(
                        f"<b>{label}</b>",
                        styles["BodyText"],
                    ),
                    Paragraph(
                        str(value or ""),
                        styles["BodyText"],
                    ),
                ]
            )

        table = Table(
            data,
            colWidths=[150, 360],
        )

        table.setStyle(
            TableStyle(
                [
                    (
                        "SPAN",
                        (0, 0),
                        (1, 0),
                    ),
                    (
                        "BACKGROUND",
                        (0, 0),
                        (1, 0),
                        colors.lightgrey,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "TOP",
                    ),
                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                ]
            )
        )

        elements.append(table)

        elements.append(Spacer(1, 12))

    # =================================================
    # WO INFORMATION
    # =================================================

    create_section(
        "WO INFORMATION",
        [
            (
                "Work Order ID",
                work_order["work_order_id"],
            ),
            (
                "Request ID",
                work_order["request_id"],
            ),
            (
                "WO Date",
                work_order["wo_date"],
            ),
            (
                "Status",
                work_order["status"],
            ),
            (
                "Priority",
                work_order["priority"],
            ),
            (
                "Maintenance Type",
                work_order["maintenance_type"],
            ),
            (
                "Requester Department",
                work_order["requester_department"],
            ),
            (
                "Maintenance Department",
                work_order["maintenance_department"],
            ),
        ],
    )

    # =================================================
    # EQUIPMENT & PROBLEM
    # =================================================

    create_section(
        "EQUIPMENT & PROBLEM",
        [
            (
                "Machine",
                work_order["machine_name"],
            ),
            (
                "Product",
                work_order["product_name"],
            ),
            (
                "Production Period",
                work_order["production_period"],
            ),
            (
                "Total Downtime (minutes)",
                work_order["total_downtime_minutes"],
            ),
            (
                "Production Achievement (%)",
                work_order["production_achievement"],
            ),
            (
                "Reject Product (pcs)",
                work_order["reject_product"],
            ),
            (
                "Problem Description",
                work_order["problem_description"],
            ),
            (
                "Production Impact",
                work_order["production_impact"],
            ),
            (
                "Requested Action",
                work_order["requested_action"],
            ),
        ],
    )

    # =================================================
    # MAINTENANCE EXECUTION
    # =================================================

    create_section(
        "MAINTENANCE EXECUTION",
        [
            (
                "Maintenance PIC",
                work_order["maintenance_pic"],
            ),
            (
                "Maintenance Start",
                work_order["maintenance_start"],
            ),
            (
                "Maintenance Finish",
                work_order["maintenance_finish"],
            ),
            (
                "Failure / Damage",
                work_order["failure_damage"],
            ),
            (
                "Root Cause",
                work_order["root_cause"],
            ),
            (
                "Repair Action",
                work_order["repair_action"],
            ),
            (
                "Spare Part Used",
                work_order["spare_part_used"],
            ),
            (
                "Spare Part Quantity",
                work_order["spare_part_quantity"],
            ),
        ],
    )

    # =================================================
    # VERIFICATION & CLOSURE
    # =================================================

    create_section(
        "VERIFICATION & CLOSURE",
        [
            (
                "Machine Status",
                work_order["machine_status"],
            ),
            (
                "Verification",
                work_order["verification"],
            ),
            (
                "Remarks",
                work_order["remarks"],
            ),
            (
                "Maintenance Supervisor",
                work_order["maintenance_supervisor"],
            ),
        ],
    )

    document.build(elements)

    output.seek(0)

    return output


# =====================================================
# DOWNLOAD WORK ORDER PDF
# =====================================================


@router.get("/work-order/download/pdf")
def download_work_order_pdf(
    request_id: str = Query(...),
):
    """
    Generate and download Maintenance Work Order
    as a PDF file.
    """

    request = get_maintenance_request(
        request_id,
    )

    if not request:

        raise HTTPException(
            status_code=404,
            detail=(f"Maintenance Request " f"'{request_id}' not found."),
        )

    work_order = build_work_order(
        request,
    )

    pdf_file = create_work_order_pdf(
        work_order,
    )

    filename = f"{work_order['work_order_id']}" "_Maintenance_Work_Order.pdf"

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={"Content-Disposition": (f'attachment; filename="{filename}"')},
    )


# =====================================================
# GET WORK ORDER DATA
# =====================================================


@router.get("/work-order")
def get_work_order(
    request_id: str = Query(...),
):
    """
    Generate Work Order data from an existing
    Maintenance Request.
    """

    request = get_maintenance_request(
        request_id,
    )

    if not request:

        raise HTTPException(
            status_code=404,
            detail=(f"Maintenance Request " f"'{request_id}' not found."),
        )

    work_order = build_work_order(
        request,
    )

    return {
        "status": "success",
        "work_order": work_order,
    }


# =====================================================
# DOWNLOAD WORK ORDER EXCEL
# =====================================================


@router.get("/work-order/download")
def download_work_order(
    request_id: str = Query(...),
):
    """
    Generate and download Maintenance Work Order
    as an Excel file.
    """

    request = get_maintenance_request(
        request_id,
    )

    if not request:

        raise HTTPException(
            status_code=404,
            detail=(f"Maintenance Request " f"'{request_id}' not found."),
        )

    work_order = build_work_order(
        request,
    )

    excel_file = create_work_order_excel(
        work_order,
    )

    filename = f"{work_order['work_order_id']}" "_Maintenance_Work_Order.xlsx"

    return StreamingResponse(
        excel_file,
        media_type=(
            "application/vnd.openxmlformats-officedocument" ".spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": (f'attachment; filename="{filename}"')},
    )
